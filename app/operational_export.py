from datetime import datetime, timezone
from io import BytesIO
import json

from flask import jsonify, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .rbac import has_permission


HEADER_FILL = PatternFill("solid", fgColor="536B45")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _iso(value):
    return value.isoformat() if value else ""


def _join(value):
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return value if value is not None else ""


def _sheet(workbook, title, headers, rows):
    ws = workbook.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, column in enumerate(ws.columns, start=1):
        width = max((len(str(cell.value or "")) for cell in column), default=10) + 2
        ws.column_dimensions[get_column_letter(index)].width = min(max(width, 12), 42)
    return ws


def init_operational_export(app, app_module):
    if app.extensions.get("aplsai_operational_export"):
        return

    @app.get("/api/admin/operational-export.xlsx")
    def operational_export():
        uid = session.get("uid")
        actor = app_module.db.session.get(app_module.User, uid) if uid else None
        if not actor:
            return jsonify(error="Non autenticato."), 401
        if getattr(actor, "active", True) is False:
            session.clear()
            return jsonify(error="Account disattivato."), 401
        if not has_permission(actor.role, "staff_manage"):
            return jsonify(error="Permesso insufficiente."), 403

        wb = Workbook()
        wb.remove(wb.active)
        generated_at = datetime.now(timezone.utc)
        _sheet(wb, "Informazioni", ["Campo", "Valore"], [
            ("Pacchetto", "APLSAI HOME - Esportazione operativa"),
            ("Generato il", generated_at.isoformat()),
            ("Origine", "Area Admin APLSAI HOME"),
            ("Regola aggiornamento", "Usare gli ID come chiave; aggiornare le righe esistenti senza duplicarle."),
            ("Classificazione", "I dati di prova sono identificati e separati dai clienti reali."),
            ("Riservatezza", "Contiene dati personali: conservare in posizione protetta."),
        ])

        profiles = {p.user_id: p for p in app_module.ClientProfile.query.all()}
        clients = app_module.User.query.filter_by(role="client").order_by(app_module.User.id.asc()).all()
        _sheet(wb, "Clienti", [
            "ID cliente", "Nome", "Email", "Telefono", "Attivo", "Creato il",
            "Tipo dato", "Archiviato il", "Stato", "Strategia", "Ultimo contatto",
            "Ultima proposta", "N. proposte",
        ], [(
            u.id, u.name, u.email, u.phone, bool(getattr(u, "active", True)), _iso(u.created_at),
            "Prova" if bool(getattr(profiles.get(u.id), "is_test", False)) else "Reale",
            _iso(getattr(profiles.get(u.id), "archived_at", None)),
            profiles[u.id].status if u.id in profiles else "",
            profiles[u.id].preferred_strategy if u.id in profiles else "",
            _iso(profiles[u.id].last_contact_at) if u.id in profiles else "",
            _iso(profiles[u.id].last_proposal_at) if u.id in profiles else "",
            profiles[u.id].proposal_count or 0 if u.id in profiles else 0,
        ) for u in clients])

        profile_rows = []
        for client in clients:
            cp = profiles.get(client.id)
            try:
                data = json.loads(cp.profile_json) if cp else {}
            except (TypeError, ValueError):
                data = {}
            zone, budget, spaces = data.get("zone") or {}, data.get("budget") or {}, data.get("spaces") or {}
            profile_rows.append((
                client.id, zone.get("main", ""), zone.get("km", ""), budget.get("ideal", ""),
                budget.get("max", ""), budget.get("flex", ""), spaces.get("sqm", ""),
                spaces.get("beds", ""), spaces.get("baths", ""), _join(data.get("must")),
                data.get("timing", ""), _join(data.get("houseTypes")), _join(data.get("purchase")),
                data.get("style", ""), json.dumps(data, ensure_ascii=False),
            ))
        _sheet(wb, "Profili abitativi", [
            "ID cliente", "Tipo dato", "Archiviato il", "Zona principale", "Distanza km",
            "Budget ideale", "Budget massimo",
            "Flessibilità %", "Metratura", "Camere", "Bagni", "Indispensabili", "Tempistica",
            "Tipologie casa", "Stato acquisto", "Stile", "Profilo JSON originale",
        ], [(
            row[0],
            "Prova" if bool(getattr(profiles.get(row[0]), "is_test", False)) else "Reale",
            _iso(getattr(profiles.get(row[0]), "archived_at", None)),
            *row[1:],
        ) for row in profile_rows])

        _sheet(wb, "Immobili", [
            "ID immobile", "Riferimento", "Tipologia", "Indirizzo", "Zona", "Prezzo", "Mq",
            "Camere", "Bagni", "Piano", "Ascensore", "Esposizione", "Spazi esterni",
            "Parcheggio", "Stato manutentivo", "Classe energetica", "Stato impianti",
            "Disponibilità", "Vincoli conosciuti", "Trasformabilità", "Interventi ipotizzati",
            "Costo lavori minimo", "Costo lavori massimo", "Mesi minimi", "Mesi massimi",
            "Affidabilità dati", "Verifica tecnica", "Fonte", "Note", "Archiviato il",
            "Creato il", "Aggiornato il",
        ], [
            (
                p.id, p.ref, p.property_type, p.address, p.zone, p.price, p.sqm, p.beds,
                p.baths, p.floor, p.elevator, p.exposure, p.outdoor_spaces, p.parking,
                p.state, p.energy_class, p.systems_status, p.availability,
                p.known_constraints, p.transformation_status, p.planned_works,
                p.renovation_cost_min, p.renovation_cost_max, p.renovation_months_min,
                p.renovation_months_max, p.data_reliability, p.technical_verification,
                p.source, p.notes, _iso(p.archived_at), _iso(p.created_at), _iso(p.updated_at),
            )
            for p in app_module.Property.query.order_by(app_module.Property.id.asc()).all()
        ])

        property_ext = app.extensions.get("aplsai_property_profiles") or {}
        PropertyRevision = property_ext.get("PropertyRevision")
        revisions = PropertyRevision.query.order_by(PropertyRevision.property_id.asc(), PropertyRevision.version.asc()).all() if PropertyRevision else []
        _sheet(wb, "Storico immobili", [
            "ID revisione", "ID immobile", "Versione", "ID autore", "Motivo", "Data", "Fotografia JSON",
        ], [(
            row.id, row.property_id, row.version, row.changed_by_user_id,
            row.change_note, _iso(row.created_at), row.snapshot_json,
        ) for row in revisions])

        opportunity_ext = app.extensions.get("aplsai_opportunities") or {}
        Opportunity = opportunity_ext.get("PropertyOpportunity")
        OpportunityRevision = opportunity_ext.get("OpportunityRevision")
        opportunity_serializer = opportunity_ext.get("opportunity_dict")
        opportunities = Opportunity.query.order_by(Opportunity.id.asc()).all() if Opportunity else []
        opportunity_data = [(row, opportunity_serializer(row, include_details=True)) for row in opportunities] if opportunity_serializer else []
        _sheet(wb, "Opportunità immobiliari", [
            "ID", "Titolo", "Tipo fonte", "Fonte", "Collegamento", "Riferimento fonte",
            "Contatto", "Recapito", "Zona", "Indirizzo", "Prezzo", "Mq", "Tipologia",
            "Stato immobile", "Disponibilità", "Documenti", "Planimetria", "Analisi",
            "Affidabilità", "Fase opportunità", "Rischi", "Potenziale", "Decisione",
            "Nota decisionale", "Motivo esclusione", "Ultimo controllo", "ID immobile",
            "Versione", "Archiviata il", "Aggiornata il", "Note",
        ], [(
            row.id, row.title, row.source_type, row.source_name, row.source_url, row.external_ref,
            row.contact_name, row.contact_details, row.zone, row.address, row.price, row.sqm,
            row.property_type, row.state, row.availability, row.documents_status,
            row.planimetry_status, row.analysis_status, row.data_reliability, row.status,
            row.risks, row.potential, row.decision, row.decision_note, row.rejection_reason,
            row.last_checked_on.isoformat() if row.last_checked_on else "", row.linked_property_id,
            row.version, _iso(row.archived_at), _iso(row.updated_at), row.notes,
        ) for row in opportunities])
        _sheet(wb, "Compatibilità opportunità", [
            "ID opportunità", "Titolo", "ID cliente", "Cliente", "Esito preliminare", "Verifiche",
        ], [(
            row.id, row.title, match.get("client_id"), match.get("client_name"),
            match.get("recommendation"),
            "; ".join(f"{item.get('criterion')}: {item.get('status')}" for item in match.get("checks") or []),
        ) for row, data in opportunity_data for match in data.get("preliminary_matches", [])])
        opportunity_revisions = OpportunityRevision.query.order_by(OpportunityRevision.opportunity_id.asc(), OpportunityRevision.version.asc()).all() if OpportunityRevision else []
        _sheet(wb, "Storico opportunità", [
            "ID revisione", "ID opportunità", "Versione", "ID autore", "Motivo", "Data", "Fotografia JSON",
        ], [(
            row.id, row.opportunity_id, row.version, row.changed_by_user_id,
            row.change_note, _iso(row.created_at), row.snapshot_json,
        ) for row in opportunity_revisions])

        scenario_ext = app.extensions.get("aplsai_scenarios") or {}
        Scenario = scenario_ext.get("PropertyScenario")
        CostItem = scenario_ext.get("ScenarioCostItem")
        ScenarioRevision = scenario_ext.get("ScenarioRevision")
        scenarios = Scenario.query.order_by(Scenario.id.asc()).all() if Scenario else []
        scenario_dict = scenario_ext.get("scenario_dict")
        _sheet(wb, "Scenari", [
            "ID", "ID immobile", "Riferimento", "ID cliente", "Nome cliente", "Nome scenario",
            "Tipo", "Stato", "Descrizione", "Mq risultanti", "Camere risultanti",
            "Bagni risultanti", "Mesi minimi", "Mesi massimi", "Ipotesi", "Vincoli",
            "Validazione tecnica", "Versione", "Totale conosciuto minimo",
            "Totale conosciuto massimo", "Voci mancanti", "Archiviato il", "Aggiornato il",
        ], [(
            row.id, row.property_id, data.get("property_ref"), row.client_id, data.get("client_name"),
            row.name, row.scenario_type, row.status, row.description, row.projected_sqm,
            row.projected_beds, row.projected_baths, row.months_min, row.months_max,
            row.assumptions, row.constraints, row.technical_validation, row.version,
            data["totals"]["known_total_min"], data["totals"]["known_total_max"],
            ", ".join(data["totals"]["missing_categories"]), _iso(row.archived_at), _iso(row.updated_at),
        ) for row in scenarios for data in [scenario_dict(row)]])
        cost_items = CostItem.query.order_by(CostItem.scenario_id.asc(), CostItem.id.asc()).all() if CostItem else []
        _sheet(wb, "Costi scenari", [
            "ID voce", "ID scenario", "Categoria", "Descrizione", "Quantità", "Unità",
            "Prezzo unitario minimo", "Prezzo unitario massimo", "Totale minimo", "Totale massimo",
            "Fonte", "Affidabilità", "Creato il",
        ], [(
            row.id, row.scenario_id, row.category, row.description, row.quantity, row.unit,
            row.unit_price_min, row.unit_price_max, row.quantity * row.unit_price_min,
            row.quantity * row.unit_price_max, row.source, row.reliability, _iso(row.created_at),
        ) for row in cost_items])
        scenario_revisions = ScenarioRevision.query.order_by(ScenarioRevision.scenario_id.asc(), ScenarioRevision.version.asc()).all() if ScenarioRevision else []
        _sheet(wb, "Storico scenari", [
            "ID revisione", "ID scenario", "Versione", "ID autore", "Motivo", "Data", "Fotografia JSON",
        ], [(
            row.id, row.scenario_id, row.version, row.changed_by_user_id,
            row.change_note, _iso(row.created_at), row.snapshot_json,
        ) for row in scenario_revisions])

        feasibility_ext = app.extensions.get("aplsai_feasibility") or {}
        Analysis = feasibility_ext.get("FeasibilityAnalysis")
        AnalysisRevision = feasibility_ext.get("FeasibilityRevision")
        analysis_dict = feasibility_ext.get("analysis_dict")
        analyses = Analysis.query.order_by(Analysis.id.asc()).all() if Analysis else []
        _sheet(wb, "Fattibilità operazioni", [
            "ID", "ID immobile", "Riferimento", "ID scenario", "Scenario", "Nome analisi",
            "Stato", "Vendita attesa", "Altri ricavi", "Capitale AP", "Finanziamento esterno",
            "Risk Budget", "Margine obiettivo %", "Durata Base mesi", "Decisione",
            "Costi conosciuti Base", "Categorie mancanti", "Versione", "Note", "Aggiornato il",
        ], [(
            row.id, row.property_id, data.get("property_ref"), row.scenario_id, data.get("scenario_name"),
            row.name, row.status, row.expected_sale_value, row.other_income, row.ap_capital,
            row.external_financing, row.risk_budget, row.target_margin_percent,
            row.base_duration_months, data["results"]["decision"], data["results"]["known_cost_base"],
            ", ".join(data["results"]["missing_categories"]), row.version, row.notes, _iso(row.updated_at),
        ) for row in analyses for data in [analysis_dict(row)]])
        _sheet(wb, "Stress test", [
            "ID analisi", "Scenario rischio", "Riduzione ricavi %", "Aumento costi %",
            "Ritardo mesi", "Durata mesi", "Ricavi", "Costi", "Risultato",
            "Margine ricavi %", "Margine costi %", "ROI capitale AP %", "Perdita massima",
            "Risk Budget", "Stato rischio", "Fabbisogno cassa", "Capitale AP esposto",
            "Fabbisogno scoperto", "Prezzo massimo acquisto",
        ], [(
            row.id, case["label"], case["revenue_reduction_percent"], case["cost_increase_percent"],
            case["delay_months"], case["duration_months"], case["revenue"], case["total_cost"],
            case["profit"], case["margin_on_revenue_percent"], case["margin_on_cost_percent"],
            case["roi_ap_percent"], case["loss"], data["results"]["risk_budget"], case["risk_status"],
            case["estimated_peak_cash_need"], case["ap_exposure"], case["funding_gap"],
            case["maximum_acquisition_price"],
        ) for row in analyses for data in [analysis_dict(row)] for case in data["results"]["cases"]])
        analysis_revisions = AnalysisRevision.query.order_by(AnalysisRevision.analysis_id.asc(), AnalysisRevision.version.asc()).all() if AnalysisRevision else []
        _sheet(wb, "Storico fattibilità", [
            "ID revisione", "ID analisi", "Versione", "ID autore", "Motivo", "Data", "Fotografia JSON",
        ], [(
            row.id, row.analysis_id, row.version, row.changed_by_user_id,
            row.change_note, _iso(row.created_at), row.snapshot_json,
        ) for row in analysis_revisions])

        cash_ext = app.extensions.get("aplsai_cashflow") or {}
        CashPlan = cash_ext.get("CashFlowPlan")
        CashMovement = cash_ext.get("CashFlowMovement")
        CashRevision = cash_ext.get("CashFlowRevision")
        cash_serializer = cash_ext.get("plan_dict")
        cash_plans = CashPlan.query.order_by(CashPlan.id.asc()).all() if CashPlan else []
        _sheet(wb, "Piani di cassa", [
            "ID", "ID analisi", "Analisi", "Immobile", "Nome piano", "Mese iniziale",
            "Cassa iniziale", "Linea aggiuntiva", "Stato", "Decisione", "Versione",
            "Costi scenario", "Costi pianificati", "Da pianificare", "Sovrapianificato",
            "Archiviato il", "Aggiornato il", "Note",
        ], [(
            row.id, row.analysis_id, data.get("analysis_name"), data.get("property_ref"), row.name,
            row.start_month, row.opening_cash, row.additional_credit_limit, row.status,
            data["results"]["decision"], row.version,
            data["results"]["reconciliation"]["scenario_cost_max"],
            data["results"]["reconciliation"]["planned_cost_max"],
            data["results"]["reconciliation"]["remaining_to_schedule"],
            data["results"]["reconciliation"]["over_scheduled"],
            _iso(row.archived_at), _iso(row.updated_at), row.notes,
        ) for row in cash_plans for data in [cash_serializer(row)]])
        cash_movements = CashMovement.query.order_by(CashMovement.plan_id.asc(), CashMovement.month_index.asc()).all() if CashMovement else []
        _sheet(wb, "Movimenti di cassa", [
            "ID", "ID piano", "Mese progressivo", "Tipo", "Categoria", "Descrizione",
            "Importo minimo", "Importo massimo", "Fonte", "Affidabilità", "Automatico",
        ], [(
            row.id, row.plan_id, row.month_index, row.movement_type, row.category, row.description,
            row.amount_min, row.amount_max, row.source, row.reliability, bool(row.system_generated),
        ) for row in cash_movements])
        _sheet(wb, "Cassa mensile", [
            "ID piano", "Mese progressivo", "Mese", "Entrate min", "Entrate max",
            "Uscite min", "Uscite max", "Saldo prudente", "Saldo massimo",
        ], [(
            row.id, month["month_index"], month["month"], month["inflow_min"], month["inflow_max"],
            month["outflow_min"], month["outflow_max"], month["balance_min"], month["balance_max"],
        ) for row in cash_plans for data in [cash_serializer(row)] for month in data["results"]["months"]])
        _sheet(wb, "Stress di cassa", [
            "ID piano", "Scenario", "Saldo minimo", "Mese picco", "Fabbisogno aggiuntivo",
            "Linea disponibile", "Copertura", "Saldo finale",
        ], [(
            row.id, case["label"], case["minimum_balance"], case["peak_month"],
            case["additional_funding_need"], case["credit_limit"], case["coverage_status"],
            case["closing_balance"],
        ) for row in cash_plans for data in [cash_serializer(row)] for case in data["results"]["stress_cases"]])
        cash_revisions = CashRevision.query.order_by(CashRevision.plan_id.asc(), CashRevision.version.asc()).all() if CashRevision else []
        _sheet(wb, "Storico cassa", [
            "ID revisione", "ID piano", "Versione", "ID autore", "Motivo", "Data", "Fotografia JSON",
        ], [(
            row.id, row.plan_id, row.version, row.changed_by_user_id,
            row.change_note, _iso(row.created_at), row.snapshot_json,
        ) for row in cash_revisions])

        portfolio_ext = app.extensions.get("aplsai_portfolio") or {}
        portfolio_serializer = portfolio_ext.get("portfolio_dict")
        portfolio = portfolio_serializer(include_history=True) if portfolio_serializer else {
            "settings": {}, "results": {"cases": [], "plans": [], "decision": "Non disponibile"}, "revisions": [],
        }
        portfolio_settings = portfolio.get("settings") or {}
        portfolio_results = portfolio.get("results") or {}
        _sheet(wb, "Portafoglio", [
            "Decisione", "Piani attivi", "Liquidità AP disponibile", "Riserva minima",
            "Tetto esposizione AP", "Massimo operazioni simultanee", "Versione limiti", "Note",
        ], [(
            portfolio_results.get("decision"), portfolio_results.get("active_plan_count", 0),
            portfolio_settings.get("available_liquidity"), portfolio_settings.get("minimum_liquidity_reserve"),
            portfolio_settings.get("max_ap_exposure"), portfolio_settings.get("max_concurrent_operations"),
            portfolio_settings.get("version", 0), portfolio_settings.get("notes", ""),
        )])
        _sheet(wb, "Stress portafoglio", [
            "Scenario", "Decisione", "Assorbimento massimo", "Esposizione AP massima",
            "Mese critico", "Operazioni simultanee massime", "Liquidità residua minima",
            "Fabbisogno massimo senza copertura", "Limiti superati", "Piani incompleti",
        ], [(
            case.get("label"), case.get("decision"), case.get("peak_cash_absorption"),
            case.get("peak_ap_exposure"), case.get("peak_month"), case.get("max_concurrent_operations"),
            case.get("minimum_remaining_liquidity"), case.get("maximum_uncovered_need"),
            ", ".join(case.get("breaches") or []), ", ".join(case.get("incomplete_plans") or []),
        ) for case in portfolio_results.get("cases", [])])
        _sheet(wb, "Portafoglio mensile", [
            "Scenario", "Mese", "Entrate", "Uscite", "Assorbimento complessivo",
            "Esposizione AP", "Credito utilizzato", "Fabbisogno senza copertura",
            "Liquidità residua", "Operazioni simultanee", "Piani attivi",
        ], [(
            case.get("label"), month.get("month"), month.get("inflows"), month.get("outflows"),
            month.get("gross_cash_absorption"), month.get("ap_exposure"), month.get("credit_used"),
            month.get("uncovered_need"), month.get("remaining_liquidity"),
            month.get("concurrent_operations"),
            ", ".join(item.get("name", "") for item in month.get("active_plans") or []),
        ) for case in portfolio_results.get("cases", []) for month in case.get("months", [])])
        _sheet(wb, "Storico portafoglio", [
            "ID revisione", "ID configurazione", "Versione", "ID autore", "Motivo", "Data", "Fotografia JSON",
        ], [(
            row.get("id"), row.get("settings_id"), row.get("version"), row.get("changed_by_user_id"),
            row.get("change_note"), row.get("created_at"), json.dumps(row.get("snapshot") or {}, ensure_ascii=False),
        ) for row in portfolio.get("revisions", [])])

        capacity_ext = app.extensions.get("aplsai_capacity") or {}
        capacity_serializer = capacity_ext.get("capacity_dict")
        capacity = capacity_serializer() if capacity_serializer else {"teams": [], "allocations": [], "results": {"months": []}}
        _sheet(wb, "Squadre operative", [
            "ID", "Nome", "Società o partner", "Specializzazione", "Responsabile",
            "Capacità mensile giornate-uomo", "Fonte", "Affidabilità", "Versione",
            "Archiviato il", "Aggiornato il", "Note",
        ], [(
            row.get("id"), row.get("name"), row.get("company"), row.get("specialty"),
            row.get("responsible"), row.get("monthly_capacity_days"), row.get("source"),
            row.get("reliability"), row.get("version"), row.get("archived_at"),
            row.get("updated_at"), row.get("notes"),
        ) for row in capacity.get("teams", [])])
        _sheet(wb, "Assegnazioni operative", [
            "ID", "ID squadra", "Squadra", "ID piano", "Operazione", "Immobile", "Mese",
            "Fase", "Giornate-uomo richieste", "Stato", "Dipendenza esterna",
            "Fonte", "Affidabilità", "Aggiornato il", "Note",
        ], [(
            row.get("id"), row.get("team_id"), row.get("team_name"), row.get("plan_id"),
            row.get("plan_name"), row.get("property_ref"), row.get("month"), row.get("phase"),
            row.get("required_worker_days"), row.get("status"), row.get("external_dependency"),
            row.get("source"), row.get("reliability"), row.get("updated_at"), row.get("notes"),
        ) for row in capacity.get("allocations", [])])
        _sheet(wb, "Capacità mensile", [
            "Mese", "ID squadra", "Squadra", "Società", "Capacità giornate-uomo",
            "Giornate-uomo impegnate", "Giornate-uomo residue", "Utilizzo %", "Stato",
        ], [(
            row.get("month"), row.get("team_id"), row.get("team_name"), row.get("company"),
            row.get("capacity_days"), row.get("used_days"), row.get("remaining_days"),
            row.get("utilization_percent"), row.get("status"),
        ) for row in (capacity.get("results") or {}).get("months", [])])
        TeamRevision = capacity_ext.get("TeamRevision")
        team_revisions = TeamRevision.query.order_by(TeamRevision.team_id.asc(), TeamRevision.version.asc()).all() if TeamRevision else []
        _sheet(wb, "Storico squadre", [
            "ID revisione", "ID squadra", "Versione", "ID autore", "Motivo", "Data", "Fotografia JSON",
        ], [(
            row.id, row.team_id, row.version, row.changed_by_user_id,
            row.change_note, _iso(row.created_at), row.snapshot_json,
        ) for row in team_revisions])

        operations_ext = app.extensions.get("aplsai_operations") or {}
        Operation = operations_ext.get("ClientOperation")
        operations = Operation.query.order_by(Operation.client_id.asc()).all() if Operation else []
        _sheet(wb, "Pratiche", [
            "ID cliente", "Tipo dato", "Archiviato il", "Fase", "Stato finanziario",
            "Priorità", "Verificato il", "Prossima azione",
            "Scadenza", "Responsabile", "Motivo blocco", "Aggiornato il",
        ], [(
            op.client_id,
            "Prova" if bool(getattr(profiles.get(op.client_id), "is_test", False)) else "Reale",
            _iso(getattr(profiles.get(op.client_id), "archived_at", None)),
            op.phase, op.financial_state, op.to_dict().get("priority", ""),
            _iso(op.financial_verified_at), op.next_action, _iso(op.next_action_due_at),
            op.assigned_to, op.blocked_reason, _iso(op.updated_at),
        ) for op in operations])

        _sheet(wb, "Trattative", ["ID", "ID cliente", "ID immobile", "Riferimento", "Fase", "Aggiornato il"], [
            (d.id, d.client_id, d.property_id, d.ref, d.stage, _iso(d.updated_at))
            for d in app_module.Deal.query.order_by(app_module.Deal.id.asc()).all()
        ])
        _sheet(wb, "Referral", ["ID", "ID cliente", "Email invitato", "Codice", "Stato", "Premio", "Creato il"], [
            (r.id, r.owner_id, r.friend_email, r.code, r.status, r.reward, _iso(r.created_at))
            for r in app_module.Referral.query.order_by(app_module.Referral.id.asc()).all()
        ])
        _sheet(wb, "Documenti", ["ID", "ID cliente", "Titolo", "Riferimento documento", "Creato il"], [
            (d.id, d.client_id, d.title, d.url, _iso(d.created_at))
            for d in app_module.Document.query.order_by(app_module.Document.id.asc()).all()
        ])
        _sheet(wb, "Aggiornamenti", ["ID", "ID cliente", "Data", "Messaggio"], [
            (u.id, u.client_id, _iso(u.created_at), u.message)
            for u in app_module.Update.query.order_by(app_module.Update.id.asc()).all()
        ])
        staff = app_module.User.query.filter(app_module.User.role.in_(["staff", "operator", "partner"])).order_by(app_module.User.id.asc()).all()
        _sheet(wb, "Collaboratori", ["ID", "Nome", "Email", "Ruolo", "Attivo", "Creato il"], [
            (u.id, u.name, u.email, "admin" if u.role == "staff" else u.role, bool(getattr(u, "active", True)), _iso(u.created_at))
            for u in staff
        ])
        Audit = operations_ext.get("AuditEvent")
        events = Audit.query.order_by(Audit.id.asc()).all() if Audit else []
        _sheet(wb, "Audit", ["ID", "ID autore", "Azione", "Tipo oggetto", "ID oggetto", "Esito", "Dettaglio", "Data"], [
            (e.id, e.actor_user_id, e.action, e.object_type, e.object_id, e.outcome, e.detail, _iso(e.created_at))
            for e in events
        ])

        audit = operations_ext.get("audit")
        if audit:
            audit(actor, "operational_export", "system", "xlsx", "Pacchetto operativo scaricato")
            app_module.db.session.commit()

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        filename = f"APLSAI_HOME_Pacchetto_Operativo_{generated_at:%Y-%m-%d_%H%M}.xlsx"
        return send_file(
            stream,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            max_age=0,
        )

    app.extensions["aplsai_operational_export"] = {"installed": True}
