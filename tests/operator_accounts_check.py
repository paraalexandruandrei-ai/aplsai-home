import os
import tempfile
import unittest
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import text
from werkzeug.security import generate_password_hash

_tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
_tmp.close()
os.environ["DATABASE_URL"] = f"sqlite:///{_tmp.name}"
os.environ["FLASK_ENV"] = "testing"
os.environ["SECRET_KEY"] = "operator-account-test-secret"
os.environ["ADMIN_EMAIL"] = "admin-accounts@example.com"
os.environ["ADMIN_PASSWORD"] = "AdminAccounts12345"

import app as app_module
from app.operations import init_operations
from app.rbac_runtime import install_runtime_rbac
from app.staff_accounts import init_staff_accounts
from app.operational_export import init_operational_export
from app.client_classification import init_client_classification
from app.property_profiles import init_property_profiles
from app.scenarios import init_scenarios
from app.feasibility import init_feasibility
from app.cashflow import init_cashflow
from app.portfolio import init_portfolio
from app.capacity import init_capacity
from app.opportunities import init_opportunities
from app.outreach import init_outreach
from app.staff_protocol import init_staff_protocol


class OperatorAccountsCheck(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = app_module.create_app()
        cls.app.config.update(TESTING=True)
        init_operations(cls.app, app_module)
        install_runtime_rbac(cls.app, app_module)
        init_staff_accounts(cls.app, app_module)
        init_operational_export(cls.app, app_module)
        init_client_classification(cls.app, app_module)
        init_property_profiles(cls.app, app_module)
        init_scenarios(cls.app, app_module)
        init_feasibility(cls.app, app_module)
        init_cashflow(cls.app, app_module)
        init_portfolio(cls.app, app_module)
        init_capacity(cls.app, app_module)
        init_opportunities(cls.app, app_module)
        init_outreach(cls.app, app_module)
        init_staff_protocol(cls.app, app_module)

        with cls.app.app_context():
            for role, email in [
                ("operator", "existing-operator@example.com"),
                ("operator", "toggle-operator@example.com"),
                ("client", "account-client@example.com"),
            ]:
                if not app_module.User.query.filter_by(email=email).first():
                    app_module.db.session.add(app_module.User(
                        role=role,
                        name=f"Test {role}",
                        email=email,
                        phone="",
                        active=True,
                        password_hash=generate_password_hash("RoleTest12345", method="scrypt"),
                    ))
            app_module.db.session.commit()

    @classmethod
    def tearDownClass(cls):
        try:
            os.unlink(_tmp.name)
        except OSError:
            pass

    def setUp(self):
        app_module._login_attempts.clear()
        with self.app.app_context():
            u = app_module.User.query.filter_by(email="toggle-operator@example.com").first()
            if u:
                u.active = True
                app_module.db.session.commit()

    def login_admin(self):
        c = self.app.test_client()
        r = c.post("/api/staff/login", json={
            "email": "admin-accounts@example.com",
            "password": "AdminAccounts12345",
        })
        self.assertEqual(r.status_code, 200, r.get_json())
        return c

    def role_client(self, email):
        with self.app.app_context():
            u = app_module.User.query.filter_by(email=email).first()
            uid = u.id
        c = self.app.test_client()
        with c.session_transaction() as s:
            s["uid"] = uid
            s["nonce"] = "role-check"
            s.permanent = True
        return c

    def test_anonymous_get_is_401(self):
        self.assertEqual(self.app.test_client().get("/api/admin/operators").status_code, 401)

    def test_operator_and_client_are_403(self):
        self.assertEqual(self.role_client("existing-operator@example.com").get("/api/admin/operators").status_code, 403)
        self.assertEqual(self.role_client("account-client@example.com").get("/api/admin/operators").status_code, 403)

    def test_admin_can_list_and_create_operator(self):
        admin = self.login_admin()
        self.assertEqual(admin.get("/api/admin/operators").status_code, 200)
        r = admin.post("/api/admin/operators", json={
            "name": "Nuovo Operatore",
            "email": "new-operator@example.com",
            "password": "NewOperator12345",
        })
        self.assertEqual(r.status_code, 201, r.get_json())
        self.assertTrue(r.get_json()["operator"]["active"])

        login = self.app.test_client().post("/api/staff/login", json={
            "email": "new-operator@example.com",
            "password": "NewOperator12345",
        })
        self.assertEqual(login.status_code, 200, login.get_json())
        self.assertEqual(login.get_json().get("role"), "operator")

    def test_admin_can_deactivate_and_reactivate_operator(self):
        admin = self.login_admin()
        with self.app.app_context():
            op_id = app_module.User.query.filter_by(email="toggle-operator@example.com").first().id

        r = admin.patch(f"/api/admin/operators/{op_id}/active", json={"active": False})
        self.assertEqual(r.status_code, 200, r.get_json())
        self.assertFalse(r.get_json()["operator"]["active"])

        denied = self.app.test_client().post("/api/staff/login", json={
            "email": "toggle-operator@example.com",
            "password": "RoleTest12345",
        })
        self.assertEqual(denied.status_code, 403, denied.get_json())

        r = admin.patch(f"/api/admin/operators/{op_id}/active", json={"active": True})
        self.assertEqual(r.status_code, 200, r.get_json())
        self.assertTrue(r.get_json()["operator"]["active"])

        allowed = self.app.test_client().post("/api/staff/login", json={
            "email": "toggle-operator@example.com",
            "password": "RoleTest12345",
        })
        self.assertEqual(allowed.status_code, 200, allowed.get_json())

    def test_deactivation_invalidates_existing_session_on_next_request(self):
        c = self.app.test_client()
        login = c.post("/api/staff/login", json={
            "email": "toggle-operator@example.com",
            "password": "RoleTest12345",
        })
        self.assertEqual(login.status_code, 200, login.get_json())
        self.assertEqual(c.get("/api/staff/dashboard").status_code, 200)

        admin = self.login_admin()
        with self.app.app_context():
            op_id = app_module.User.query.filter_by(email="toggle-operator@example.com").first().id
        self.assertEqual(admin.patch(f"/api/admin/operators/{op_id}/active", json={"active": False}).status_code, 200)
        self.assertEqual(c.get("/api/staff/dashboard").status_code, 401)

    def test_deactivation_is_audited(self):
        admin = self.login_admin()
        with self.app.app_context():
            op_id = app_module.User.query.filter_by(email="toggle-operator@example.com").first().id
        self.assertEqual(admin.patch(f"/api/admin/operators/{op_id}/active", json={"active": False}).status_code, 200)
        audit = admin.get("/api/staff/audit?limit=50")
        self.assertEqual(audit.status_code, 200)
        self.assertTrue(any(
            e.get("action") == "operator_deactivate" and e.get("object_id") == str(op_id)
            for e in audit.get_json().get("events", [])
        ))

    def test_duplicate_email_is_rejected(self):
        admin = self.login_admin()
        r = admin.post("/api/admin/operators", json={
            "name": "Duplicato",
            "email": "existing-operator@example.com",
            "password": "Duplicate12345",
        })
        self.assertEqual(r.status_code, 409)

    def test_weak_password_is_rejected(self):
        admin = self.login_admin()
        r = admin.post("/api/admin/operators", json={
            "name": "Debole",
            "email": "weak-operator@example.com",
            "password": "debole",
        })
        self.assertEqual(r.status_code, 400)

    def test_creation_is_audited(self):
        admin = self.login_admin()
        r = admin.post("/api/admin/operators", json={
            "name": "Audit Operatore",
            "email": "audit-operator@example.com",
            "password": "AuditOperator12345",
        })
        self.assertEqual(r.status_code, 201, r.get_json())
        operator_id = str(r.get_json()["operator"]["id"])
        audit = admin.get("/api/staff/audit?limit=50")
        self.assertEqual(audit.status_code, 200)
        self.assertTrue(any(
            e.get("action") == "operator_create" and e.get("object_id") == operator_id
            for e in audit.get_json().get("events", [])
        ))

    def test_staff_identity_and_password_change(self):
        admin = self.login_admin()
        me = admin.get("/api/staff/me")
        self.assertEqual(me.status_code, 200, me.get_json())
        self.assertEqual(me.get_json()["staff"]["role"], "admin")

        wrong = admin.post("/api/staff/password", json={
            "current_password": "WrongPassword1",
            "new_password": "ChangedAdmin12345",
        })
        self.assertEqual(wrong.status_code, 401)

        changed = admin.post("/api/staff/password", json={
            "current_password": "AdminAccounts12345",
            "new_password": "ChangedAdmin12345",
        })
        self.assertEqual(changed.status_code, 200, changed.get_json())

        old_login = self.app.test_client().post("/api/staff/login", json={
            "email": "admin-accounts@example.com",
            "password": "AdminAccounts12345",
        })
        self.assertEqual(old_login.status_code, 401)
        new_login = self.app.test_client().post("/api/staff/login", json={
            "email": "admin-accounts@example.com",
            "password": "ChangedAdmin12345",
        })
        self.assertEqual(new_login.status_code, 200, new_login.get_json())

        with self.app.app_context():
            u = app_module.User.query.filter_by(email="admin-accounts@example.com").first()
            u.password_hash = generate_password_hash("AdminAccounts12345", method="scrypt")
            app_module.db.session.commit()

    def test_admin_password_recovery_runs_once(self):
        with self.app.app_context():
            admin = app_module.User.query.filter_by(email="admin-accounts@example.com").first()
            admin.password_hash = generate_password_hash("StaleAdmin12345", method="scrypt")
            app_module.db.session.execute(
                text("DELETE FROM aplsai_schema_migration WHERE id=:id"),
                {"id": app_module.ADMIN_PASSWORD_RECOVERY_MIGRATION},
            )
            app_module.db.session.commit()
            app_module.seed_admin()

        recovered = self.login_admin()
        changed = recovered.post("/api/staff/password", json={
            "current_password": "AdminAccounts12345",
            "new_password": "RecoveredAdmin12345",
        })
        self.assertEqual(changed.status_code, 200, changed.get_json())

        with self.app.app_context():
            app_module.seed_admin()

        survived = self.app.test_client().post("/api/staff/login", json={
            "email": "admin-accounts@example.com",
            "password": "RecoveredAdmin12345",
        })
        self.assertEqual(survived.status_code, 200, survived.get_json())

        with self.app.app_context():
            admin = app_module.User.query.filter_by(email="admin-accounts@example.com").first()
            admin.password_hash = generate_password_hash("AdminAccounts12345", method="scrypt")
            app_module.db.session.commit()

    def test_operator_can_change_own_password_but_not_manage_operators(self):
        operator = self.role_client("existing-operator@example.com")
        me = operator.get("/api/staff/me")
        self.assertEqual(me.status_code, 200, me.get_json())
        self.assertEqual(me.get_json()["staff"]["role"], "operator")
        self.assertEqual(operator.get("/api/admin/operators").status_code, 403)

        changed = operator.post("/api/staff/password", json={
            "current_password": "RoleTest12345",
            "new_password": "OperatorChanged12345",
        })
        self.assertEqual(changed.status_code, 200, changed.get_json())
        with self.app.app_context():
            u = app_module.User.query.filter_by(email="existing-operator@example.com").first()
            u.password_hash = generate_password_hash("RoleTest12345", method="scrypt")
            app_module.db.session.commit()

    def test_admin_can_download_operational_excel(self):
        admin = self.login_admin()
        response = admin.get("/api/admin/operational-export.xlsx")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data[:2], b"PK")
        self.assertIn("application/vnd.openxmlformats", response.content_type)
        workbook = load_workbook(BytesIO(response.data), read_only=True)
        self.assertIn("Immobili", workbook.sheetnames)
        self.assertIn("Storico immobili", workbook.sheetnames)
        self.assertIn("Scenari", workbook.sheetnames)
        self.assertIn("Costi scenari", workbook.sheetnames)
        self.assertIn("Storico scenari", workbook.sheetnames)
        self.assertIn("Fattibilità operazioni", workbook.sheetnames)
        self.assertIn("Stress test", workbook.sheetnames)
        self.assertIn("Storico fattibilità", workbook.sheetnames)
        self.assertIn("Piani di cassa", workbook.sheetnames)
        self.assertIn("Movimenti di cassa", workbook.sheetnames)
        self.assertIn("Cassa mensile", workbook.sheetnames)
        self.assertIn("Stress di cassa", workbook.sheetnames)
        self.assertIn("Storico cassa", workbook.sheetnames)
        self.assertIn("Portafoglio", workbook.sheetnames)
        self.assertIn("Stress portafoglio", workbook.sheetnames)
        self.assertIn("Portafoglio mensile", workbook.sheetnames)
        self.assertIn("Storico portafoglio", workbook.sheetnames)
        self.assertIn("Squadre operative", workbook.sheetnames)
        self.assertIn("Assegnazioni operative", workbook.sheetnames)
        self.assertIn("Capacità mensile", workbook.sheetnames)
        self.assertIn("Storico squadre", workbook.sheetnames)
        self.assertIn("Opportunità immobiliari", workbook.sheetnames)
        self.assertIn("Compatibilità opportunità", workbook.sheetnames)
        self.assertIn("Storico opportunità", workbook.sheetnames)
        self.assertIn("Richieste informazioni", workbook.sheetnames)
        self.assertIn("Risposte immobili", workbook.sheetnames)
        self.assertIn("Regole staff", workbook.sheetnames)
        self.assertIn("Prese visione staff", workbook.sheetnames)
        headers = [cell.value for cell in next(workbook["Immobili"].iter_rows(max_row=1))]
        self.assertIn("Trasformabilità", headers)
        self.assertIn("Costo lavori minimo", headers)

        operator = self.role_client("existing-operator@example.com")
        self.assertEqual(operator.get("/api/admin/operational-export.xlsx").status_code, 403)

    def test_advanced_property_profile_history_validation_and_archive(self):
        registration = self.app.test_client().post("/api/register", json={
            "name": "Cliente Matching Avanzato",
            "email": "advanced-match-client@example.com",
            "phone": "+393331112233",
            "password": "AdvancedMatch12345",
            "profile": {
                "zone": {"main": "Roma", "km": 15},
                "budget": {"ideal": 290000, "max": 320000, "flex": 5},
                "spaces": {"sqm": 80, "beds": 2, "baths": 1},
                "timing": "6-12 mesi", "style": "Moderno essenziale",
                "must": ["Ascensore"], "houseTypes": ["Appartamento"],
                "purchase": ["Da ristrutturare"],
            },
        })
        self.assertEqual(registration.status_code, 201, registration.get_json())
        client_id = registration.get_json()["client"]["id"]
        admin = self.login_admin()
        created = admin.post("/api/staff/properties", json={
            "ref": "IMM-PROFILO-TEST",
            "property_type": "Appartamento",
            "address": "Da verificare",
            "zone": "Roma",
            "price": 240000,
            "sqm": 82,
            "beds": 2,
            "baths": 1,
            "state": "Da ristrutturare",
            "availability": "Da verificare",
            "elevator": None,
            "energy_class": "Da verificare",
            "systems_status": "Da verificare",
            "transformation_status": "Verifica in corso",
            "planned_works": "Riduzione rischi e ridistribuzione interna da validare.",
            "renovation_cost_min": 45000,
            "renovation_cost_max": 70000,
            "renovation_months_min": 4,
            "renovation_months_max": 7,
            "data_reliability": "Dichiarato",
            "technical_verification": "Da verificare",
        })
        self.assertEqual(created.status_code, 201, created.get_json())
        prop = created.get_json()["property"]
        property_id = prop["id"]
        self.assertEqual(prop["transformation_status"], "Verifica in corso")
        self.assertEqual(prop["renovation_cost_min"], 45000)

        invalid = admin.patch(f"/api/staff/properties/{property_id}", json={
            "renovation_cost_min": 80000,
            "renovation_cost_max": 60000,
        })
        self.assertEqual(invalid.status_code, 400, invalid.get_json())

        updated = admin.patch(f"/api/staff/properties/{property_id}", json={
            "transformation_status": "Trasformabile",
            "technical_verification": "Verificato",
            "data_reliability": "Verificato",
            "renovation_cost_min": 50000,
            "renovation_cost_max": 68000,
            "change_note": "Validazione tecnica completata",
        })
        self.assertEqual(updated.status_code, 200, updated.get_json())
        self.assertEqual(updated.get_json()["property"]["transformation_status"], "Trasformabile")

        matching = admin.get(f"/api/staff/match/client/{client_id}")
        self.assertEqual(matching.status_code, 200, matching.get_json())
        matched = next(row for row in matching.get_json()["results"] if row["property"]["id"] == property_id)
        self.assertIn("score_current", matched)
        self.assertIn("score_potential", matched)
        self.assertEqual(matched["engine_version"], "APL-MATCH-2.0")
        self.assertEqual(sum(row["weight"] for row in matched["criteria"]), 100)
        self.assertIn("excluded_items", matched["economics"])

        scenario_created = admin.post("/api/staff/scenarios", json={
            "property_id": property_id, "client_id": client_id,
            "name": "Scenario famiglia equilibrato", "scenario_type": "Equilibrato",
            "status": "Da verificare", "description": "Ridistribuzione interna da validare.",
            "projected_sqm": 82, "projected_beds": 2, "projected_baths": 2,
            "months_min": 4, "months_max": 7,
            "assumptions": "Misure da rilievo preliminare.",
            "constraints": "Verifica urbanistica necessaria.",
            "technical_validation": "Da verificare",
        })
        self.assertEqual(scenario_created.status_code, 201, scenario_created.get_json())
        scenario_id = scenario_created.get_json()["scenario"]["id"]
        self.assertEqual(scenario_created.get_json()["scenario"]["version"], 1)

        invalid_item = admin.post(f"/api/staff/scenarios/{scenario_id}/cost-items", json={
            "category": "Finiture", "description": "Pavimenti", "quantity": 82,
            "unit": "m²", "unit_price_min": 80, "unit_price_max": 60,
        })
        self.assertEqual(invalid_item.status_code, 400, invalid_item.get_json())

        item_added = admin.post(f"/api/staff/scenarios/{scenario_id}/cost-items", json={
            "category": "Demolizioni e opere edili", "description": "Opere interne",
            "quantity": 82, "unit": "m²", "unit_price_min": 500,
            "unit_price_max": 650, "source": "Preventivo preliminare",
            "reliability": "Dichiarato",
        })
        self.assertEqual(item_added.status_code, 201, item_added.get_json())
        scenario = item_added.get_json()["scenario"]
        self.assertEqual(scenario["version"], 2)
        self.assertEqual(scenario["totals"]["known_total_min"], 281000)
        self.assertEqual(scenario["totals"]["known_total_max"], 293300)
        self.assertFalse(scenario["totals"]["complete"])

        scenario_updated = admin.patch(f"/api/staff/scenarios/{scenario_id}", json={
            "status": "Validato dal tecnico", "technical_validation": "Verificato",
            "change_note": "Validazione distributiva completata",
        })
        self.assertEqual(scenario_updated.status_code, 200, scenario_updated.get_json())
        self.assertEqual(scenario_updated.get_json()["scenario"]["version"], 3)
        scenario_detail = admin.get(f"/api/staff/scenarios/{scenario_id}")
        self.assertEqual(scenario_detail.status_code, 200, scenario_detail.get_json())
        self.assertEqual([row["version"] for row in scenario_detail.get_json()["scenario"]["revisions"][:3]], [3, 2, 1])

        for category in ("Tecnici e pratiche", "Imposte, notaio e agenzia", "Costi finanziari", "Fondo imprevisti"):
            added = admin.post(f"/api/staff/scenarios/{scenario_id}/cost-items", json={
                "category": category, "description": category, "quantity": 1,
                "unit": "corpo", "unit_price_min": 1000, "unit_price_max": 1000,
                "source": "Ipotesi test", "reliability": "Dichiarato",
            })
            self.assertEqual(added.status_code, 201, added.get_json())

        assumptions = {
            "base": {"revenue_reduction_percent": 0, "cost_increase_percent": 0, "delay_months": 0},
            "prudente": {"revenue_reduction_percent": 5, "cost_increase_percent": 5, "delay_months": 2},
            "stress": {"revenue_reduction_percent": 10, "cost_increase_percent": 10, "delay_months": 4},
            "doppio_stress": {"revenue_reduction_percent": 20, "cost_increase_percent": 20, "delay_months": 8},
        }
        analysis_created = admin.post("/api/staff/feasibility", json={
            "property_id": property_id, "scenario_id": scenario_id,
            "name": "Tenuta operazione", "status": "Da verificare",
            "expected_sale_value": 340000, "other_income": 0,
            "ap_capital": 100000, "external_financing": 100000,
            "risk_budget": 20000, "target_margin_percent": 10,
            "base_duration_months": 7, "assumptions": assumptions,
            "notes": "Valori di prova per collaudo.",
        })
        self.assertEqual(analysis_created.status_code, 201, analysis_created.get_json())
        analysis = analysis_created.get_json()["analysis"]
        analysis_id = analysis["id"]
        self.assertEqual(analysis["results"]["decision"], "NO-GO / RISTRUTTURARE")
        self.assertEqual(analysis["results"]["known_cost_base"], 297300)
        double_stress = next(row for row in analysis["results"]["cases"] if row["key"] == "doppio_stress")
        self.assertEqual(double_stress["profit"], -36760)
        self.assertEqual(double_stress["risk_status"], "superato")
        self.assertEqual(double_stress["maximum_acquisition_price"], 176040)

        invalid_assumptions = dict(assumptions)
        invalid_assumptions["stress"] = {"revenue_reduction_percent": 2, "cost_increase_percent": 2, "delay_months": 1}
        invalid_analysis = admin.patch(f"/api/staff/feasibility/{analysis_id}", json={"assumptions": invalid_assumptions})
        self.assertEqual(invalid_analysis.status_code, 400, invalid_analysis.get_json())

        analysis_updated = admin.patch(f"/api/staff/feasibility/{analysis_id}", json={
            "expected_sale_value": 345000, "change_note": "Aggiornato valore di uscita",
        })
        self.assertEqual(analysis_updated.status_code, 200, analysis_updated.get_json())
        self.assertEqual(analysis_updated.get_json()["analysis"]["version"], 2)
        analysis_detail = admin.get(f"/api/staff/feasibility/{analysis_id}")
        self.assertEqual([row["version"] for row in analysis_detail.get_json()["analysis"]["revisions"][:2]], [2, 1])

        cash_created = admin.post("/api/staff/cashflow", json={
            "analysis_id": analysis_id, "name": "Cassa operazione test",
            "start_month": "2026-09", "opening_cash": 100000,
            "additional_credit_limit": 110000, "status": "Da verificare",
            "notes": "Calendario di collaudo.",
        })
        self.assertEqual(cash_created.status_code, 201, cash_created.get_json())
        cash_plan = cash_created.get_json()["plan"]
        cash_plan_id = cash_plan["id"]
        self.assertEqual(cash_plan["results"]["decision"], "PIANO DA COMPLETARE")
        self.assertEqual(len([m for m in cash_plan["movements"] if m["system_generated"]]), 3)

        movement_added = admin.post(f"/api/staff/cashflow/{cash_plan_id}/movements", json={
            "month_index": 2, "movement_type": "Uscita", "category": "Lavori",
            "description": "SAL complessivi pianificati", "amount_min": 50000,
            "amount_max": 57300, "source": "Scenario analitico", "reliability": "Documentato",
        })
        self.assertEqual(movement_added.status_code, 201, movement_added.get_json())
        cash_plan = movement_added.get_json()["plan"]
        self.assertEqual(cash_plan["results"]["reconciliation"]["status"], "riconciliato")
        self.assertEqual(cash_plan["results"]["decision"], "COPERTURA ADEGUATA")
        cash_double = next(row for row in cash_plan["results"]["stress_cases"] if row["key"] == "doppio_stress")
        self.assertEqual(cash_double["minimum_balance"], -108760)
        self.assertEqual(cash_double["peak_month"], "Novembre 2026")
        self.assertEqual(cash_double["coverage_status"], "coperto")

        automatic_movement = next(m for m in cash_plan["movements"] if m["system_generated"])
        protected = admin.delete(f"/api/staff/cashflow/{cash_plan_id}/movements/{automatic_movement['id']}", json={})
        self.assertEqual(protected.status_code, 409, protected.get_json())

        cash_updated = admin.patch(f"/api/staff/cashflow/{cash_plan_id}", json={
            "additional_credit_limit": 120000, "change_note": "Aumentata linea disponibile",
        })
        self.assertEqual(cash_updated.status_code, 200, cash_updated.get_json())
        self.assertEqual(cash_updated.get_json()["plan"]["version"], 3)
        cash_detail = admin.get(f"/api/staff/cashflow/{cash_plan_id}")
        self.assertEqual([row["version"] for row in cash_detail.get_json()["plan"]["revisions"][:3]], [3, 2, 1])

        portfolio_created = admin.patch("/api/staff/portfolio/settings", json={
            "available_liquidity": 120000, "minimum_liquidity_reserve": 10000,
            "max_ap_exposure": 100000, "max_concurrent_operations": 1,
            "notes": "Limiti di prova deliberati per il collaudo.",
            "change_note": "Configurazione iniziale di collaudo",
        })
        self.assertEqual(portfolio_created.status_code, 200, portfolio_created.get_json())
        portfolio = portfolio_created.get_json()["portfolio"]
        self.assertEqual(portfolio["results"]["decision"], "SOSTENIBILE")
        self.assertEqual(portfolio["results"]["active_plan_count"], 1)
        portfolio_double = next(row for row in portfolio["results"]["cases"] if row["key"] == "doppio_stress")
        self.assertEqual(portfolio_double["peak_cash_absorption"], 208760)
        self.assertEqual(portfolio_double["peak_ap_exposure"], 100000)
        self.assertEqual(portfolio_double["peak_month"], "Novembre 2026")
        self.assertEqual(portfolio_double["minimum_remaining_liquidity"], 20000)
        self.assertEqual(portfolio_double["maximum_uncovered_need"], 0)
        self.assertEqual(portfolio_double["max_concurrent_operations"], 1)

        operator = self.role_client("existing-operator@example.com")
        self.assertEqual(operator.get("/api/staff/portfolio").status_code, 200)
        self.assertEqual(operator.patch("/api/staff/portfolio/settings", json={
            "available_liquidity": 999999,
        }).status_code, 403)

        portfolio_updated = admin.patch("/api/staff/portfolio/settings", json={
            "available_liquidity": 120000, "minimum_liquidity_reserve": 10000,
            "max_ap_exposure": 90000, "max_concurrent_operations": 1,
            "notes": "Limite più prudente.", "change_note": "Ridotto tetto esposizione",
        })
        self.assertEqual(portfolio_updated.status_code, 200, portfolio_updated.get_json())
        portfolio = portfolio_updated.get_json()["portfolio"]
        self.assertEqual(portfolio["results"]["decision"], "BLOCCO / RIPIANIFICARE")
        self.assertIn("Tetto di esposizione AP superato", next(
            row for row in portfolio["results"]["cases"] if row["key"] == "doppio_stress"
        )["breaches"])
        self.assertEqual([row["version"] for row in portfolio["revisions"][:2]], [2, 1])

        team_created = admin.post("/api/staff/capacity/teams", json={
            "name": "Squadra collaudo", "company": "NP Costruzioni",
            "specialty": "Ristrutturazioni interne", "responsible": "Responsabile prova",
            "monthly_capacity_days": 100, "source": "Disponibilità di collaudo",
            "reliability": "Documentato", "notes": "Dati esclusivamente di test.",
        })
        self.assertEqual(team_created.status_code, 201, team_created.get_json())
        team_id = team_created.get_json()["team"]["id"]

        allocation_created = admin.post("/api/staff/capacity/allocations", json={
            "team_id": team_id, "plan_id": cash_plan_id, "month": "2026-11",
            "phase": "Lavori interni", "required_worker_days": 120,
            "status": "Confermato", "external_dependency": "Nessuna",
            "source": "Cronoprogramma di prova", "reliability": "Documentato",
        })
        self.assertEqual(allocation_created.status_code, 201, allocation_created.get_json())
        capacity = allocation_created.get_json()["capacity"]
        self.assertEqual(capacity["results"]["decision"], "SOVRACCARICO / RIPIANIFICARE")
        self.assertEqual(capacity["results"]["overloaded_month_count"], 1)
        month = capacity["results"]["months"][0]
        self.assertEqual(month["used_days"], 120)
        self.assertEqual(month["remaining_days"], -20)

        team_updated = admin.patch(f"/api/staff/capacity/teams/{team_id}", json={
            "monthly_capacity_days": 150, "change_note": "Aumentata capacità documentata",
        })
        self.assertEqual(team_updated.status_code, 200, team_updated.get_json())
        capacity = team_updated.get_json()["capacity"]
        self.assertEqual(capacity["results"]["decision"], "CAPACITÀ DISPONIBILE")
        self.assertEqual(capacity["results"]["months"][0]["remaining_days"], 30)
        team_detail = admin.get(f"/api/staff/capacity/teams/{team_id}")
        self.assertEqual([row["version"] for row in team_detail.get_json()["team"]["revisions"][:2]], [2, 1])

        operator_capacity = operator.get("/api/staff/capacity")
        self.assertEqual(operator_capacity.status_code, 200, operator_capacity.get_json())

        detail = admin.get(f"/api/staff/properties/{property_id}")
        self.assertEqual(detail.status_code, 200, detail.get_json())
        revisions = detail.get_json()["revisions"]
        self.assertEqual([row["version"] for row in revisions[:2]], [2, 1])
        self.assertEqual(revisions[0]["change_note"], "Validazione tecnica completata")

        archived = admin.patch(f"/api/staff/properties/{property_id}/archive", json={"archived": True})
        self.assertEqual(archived.status_code, 200, archived.get_json())
        self.assertIsNotNone(archived.get_json()["property"]["archived_at"])
        self.assertEqual(admin.get(f"/api/staff/match/property/{property_id}").status_code, 409)

        restored = admin.patch(f"/api/staff/properties/{property_id}/archive", json={"archived": False})
        self.assertEqual(restored.status_code, 200, restored.get_json())
        self.assertIsNone(restored.get_json()["property"]["archived_at"])

        admin.patch(f"/api/admin/clients/{client_id}/classification", json={"is_test": True, "archived": False})
        deleted = admin.delete(f"/api/admin/clients/{client_id}/test-data", json={"confirm": "ELIMINA"})
        self.assertEqual(deleted.status_code, 200, deleted.get_json())
        self.assertEqual(admin.get(f"/api/staff/scenarios/{scenario_id}").status_code, 404)
        self.assertEqual(admin.get(f"/api/staff/feasibility/{analysis_id}").status_code, 404)
        self.assertEqual(admin.get(f"/api/staff/cashflow/{cash_plan_id}").status_code, 404)
        self.assertEqual(admin.get("/api/staff/capacity").get_json()["capacity"]["results"]["active_allocation_count"], 0)

    def test_admin_can_separate_test_clients_and_archive_without_deleting(self):
        profile = {
            "zone": {"main": "Roma", "km": 10},
            "budget": {"ideal": 250000, "max": 300000, "flex": 5},
            "spaces": {"sqm": 75, "beds": 2, "baths": 1},
            "timing": "3-6 mesi", "style": "Moderno",
            "must": ["Balcone"], "houseTypes": ["Appartamento"],
            "purchase": ["Mutuo"],
        }
        registration = self.app.test_client().post("/api/register", json={
            "name": "Cliente Classificazione",
            "email": "classification-client@example.com",
            "phone": "+393331234567",
            "password": "Classification12345",
            "profile": profile,
        })
        self.assertEqual(registration.status_code, 201, registration.get_json())
        client_id = registration.get_json()["client"]["id"]
        self.assertFalse(registration.get_json()["client"]["is_test"])

        admin = self.login_admin()
        protected = admin.delete(f"/api/admin/clients/{client_id}/test-data", json={
            "confirm": "ELIMINA",
        })
        self.assertEqual(protected.status_code, 409)

        marked = admin.patch(f"/api/admin/clients/{client_id}/classification", json={
            "is_test": True, "archived": False,
        })
        self.assertEqual(marked.status_code, 200, marked.get_json())
        self.assertTrue(marked.get_json()["client"]["is_test"])

        operations = admin.get("/api/staff/operations").get_json()["results"]
        self.assertNotIn(client_id, [row["client"]["id"] for row in operations])

        archived = admin.patch(f"/api/admin/clients/{client_id}/classification", json={
            "is_test": True, "archived": True,
        })
        self.assertEqual(archived.status_code, 200, archived.get_json())
        self.assertIsNotNone(archived.get_json()["client"]["archived_at"])

        operator = self.role_client("existing-operator@example.com")
        self.assertEqual(operator.patch(
            f"/api/admin/clients/{client_id}/classification",
            json={"is_test": False, "archived": False},
        ).status_code, 403)

        deleted = admin.delete(f"/api/admin/clients/{client_id}/test-data", json={
            "confirm": "ELIMINA",
        })
        self.assertEqual(deleted.status_code, 200, deleted.get_json())
        with self.app.app_context():
            self.assertIsNone(app_module.db.session.get(app_module.User, client_id))

    def test_opportunity_inbox_duplicate_match_and_promotion_gate(self):
        registration = self.app.test_client().post("/api/register", json={
            "name": "Cliente Opportunità", "email": "opportunity-client@example.com",
            "phone": "+393339998877", "password": "Opportunity12345",
            "profile": {
                "zone": {"main": "Roma Nord", "km": 15},
                "budget": {"ideal": 260000, "max": 300000, "flex": 5},
                "spaces": {"sqm": 70, "beds": 2, "baths": 1},
                "timing": "3-6 mesi", "style": "Moderno",
                "must": [], "houseTypes": ["Appartamento"], "purchase": ["Da ristrutturare"],
            },
        })
        self.assertEqual(registration.status_code, 201, registration.get_json())
        client_id = registration.get_json()["client"]["id"]
        admin = self.login_admin()
        payload = {
            "title": "Annuncio Roma Nord", "source_type": "Portale",
            "source_name": "Fonte di prova", "source_url": "https://example.com/annuncio/123/?utm_source=test",
            "external_ref": "ANN-123", "zone": "Roma Nord", "address": "Via di prova 1",
            "price": 250000, "sqm": 75, "property_type": "Appartamento",
            "state": "Da ristrutturare", "availability": "Disponibile",
            "last_checked_on": "2026-09-05", "status": "Nuova",
            "documents_status": "Da verificare", "planimetry_status": "Da verificare",
            "analysis_status": "Non iniziata", "data_reliability": "Dichiarato",
            "decision": "Da decidere", "notes": "Dati esclusivamente di collaudo.",
        }
        created = admin.post("/api/staff/opportunities", json=payload)
        self.assertEqual(created.status_code, 201, created.get_json())
        opportunity = created.get_json()["opportunity"]
        opportunity_id = opportunity["id"]
        preliminary = next(row for row in opportunity["preliminary_matches"] if row["client_id"] == client_id)
        self.assertEqual(preliminary["recommendation"], "COMPATIBILITÀ PRELIMINARE")

        duplicate = dict(payload)
        duplicate["source_url"] = "https://example.com/annuncio/123?utm_campaign=copy"
        duplicate["external_ref"] = "ANN-124"
        self.assertEqual(admin.post("/api/staff/opportunities", json=duplicate).status_code, 409)

        blocked = admin.post(f"/api/staff/opportunities/{opportunity_id}/promote", json={})
        self.assertEqual(blocked.status_code, 409, blocked.get_json())
        self.assertIn("analisi preliminare", blocked.get_json()["missing"])

        invalid_decision = admin.patch(f"/api/staff/opportunities/{opportunity_id}", json={
            "decision": "Non procedere", "rejection_reason": "",
        })
        self.assertEqual(invalid_decision.status_code, 400, invalid_decision.get_json())

        updated = admin.patch(f"/api/staff/opportunities/{opportunity_id}", json={
            "status": "Analisi tecnica", "documents_status": "Parziali",
            "planimetry_status": "Disponibile", "analysis_status": "Preliminare completata",
            "decision": "Procedere", "decision_note": "Compatibile con domanda e budget; proseguire con verifiche complete.",
            "change_note": "Controllo preliminare completato",
        })
        self.assertEqual(updated.status_code, 200, updated.get_json())
        self.assertEqual(updated.get_json()["opportunity"]["version"], 2)

        promoted = admin.post(f"/api/staff/opportunities/{opportunity_id}/promote", json={})
        self.assertEqual(promoted.status_code, 201, promoted.get_json())
        property_id = promoted.get_json()["property"]["id"]
        self.assertEqual(promoted.get_json()["property"]["ref"], f"OPP-{opportunity_id:05d}")
        self.assertEqual(promoted.get_json()["property"]["technical_verification"], "Verifica in corso")
        self.assertEqual(promoted.get_json()["opportunity"]["linked_property_id"], property_id)
        self.assertEqual([row["version"] for row in promoted.get_json()["opportunity"]["revisions"][:3]], [3, 2, 1])
        self.assertEqual(admin.post(f"/api/staff/opportunities/{opportunity_id}/promote", json={}).status_code, 409)

        property_matches = admin.get(f"/api/staff/match/property/{property_id}")
        self.assertEqual(property_matches.status_code, 200, property_matches.get_json())
        self.assertIn(client_id, [row["client"]["id"] for row in property_matches.get_json()["results"]])


if __name__ == "__main__":
    unittest.main()
